from openpyxl import Workbook, utils
from db_func import select_data, select_data_dict, select_data_single, insert_many
from variables import ttl_ratiosum_process
# import datetime
from datetime import datetime, timedelta
# Получаем сегодняшнюю дату
today = datetime.now()
# Находим номер сегодняшнего дня недели (0 - понедельник, 6 - воскресенье)
today_weekday = today.weekday()
# Вычисляем дату понедельника текущей недели
monday = today - timedelta(days=today_weekday)
# Вычисляем дату воскресенья текущей недели
sunday = monday + timedelta(days=6)

#print(bonus_ratio_editor(0.5, 10))
def bonus_ratio_editor(bonus, my_num):
    if bonus<1 and bonus>0:
        my_num = my_num*bonus
    elif bonus>1:
        my_num = bonus
    elif bonus==1 or bonus==0:
        my_num=0
    return my_num

def womans_report(usernames, points):
    sqldata1 = select_data_dict("""
    SELECT 
    u.name,
    uw.telega_id, 
    p.point_name,
    (SELECT SUM(point_score) FROM userpoints_weekly WHERE telega_id = uw.telega_id AND point_name = uw.point_name GROUP BY point_name, telega_id)  AS totalsum_point_score,
    SUM(uw.point_score) as point_week_sum,
    p.ratio, 
    p.mins,
    (SELECT bonus_ratio FROM bonus WHERE bonus_name = 'Бонус минимумов') as bonus_of_mins,                      
    (SELECT CAST(SUM(bonusmins_elment) AS FLOAT) / (SELECT COUNT(point_name) FROM points) FROM (SELECT userpoints_weekly.telega_id, userpoints_weekly.point_name, COUNT(points.point_name) AS point_count, CASE WHEN SUM(userpoints_weekly.point_score) / points.mins > 1 THEN 1 ELSE 0 END AS bonusmins_elment FROM userpoints_weekly LEFT JOIN points ON userpoints_weekly.point_name = points.point_name WHERE date BETWEEN ? AND ? GROUP BY userpoints_weekly.point_name, userpoints_weekly.telega_id) as bonus_of_mins_table WHERE bonus_of_mins_table.telega_id = uw.telega_id GROUP BY telega_id) AS bonus_of_mins_ratio,                                              
    (SELECT SUM(CASE WHEN count_date=7 THEN 1 ELSE 0 END) FROM (SELECT telega_id, COUNT(DISTINCT date) AS count_date FROM userpoints_weekly WHERE telega_id = uw.telega_id AND date BETWEEN ? AND ? GROUP BY point_name) AS date_sum GROUP BY telega_id) AS count_row_bonus,                       
    (SELECT bonus_ratio FROM bonus WHERE bonus_name = 'Бонус строки') as bonus_of_rows,
    (SELECT SUM(column_bonus) AS total_bonus FROM (SELECT usrw.telega_id, usrw.date, CASE WHEN (SELECT COUNT(DISTINCT point_name) FROM userpoints_weekly WHERE telega_id = usrw.telega_id AND date = usrw.date) = (SELECT COUNT(point_name) FROM points) THEN 1 ELSE 0 END AS column_bonus FROM userpoints_weekly usrw WHERE date BETWEEN ? AND ? GROUP BY usrw.date, usrw.telega_id) AS bonuses WHERE telega_id = uw.telega_id GROUP BY telega_id) AS count_column_bonus,
    (SELECT bonus_ratio FROM bonus WHERE bonus_name = 'Бонус столбца') as bonus_of_cols,
    (SELECT SUM(point_score*ratio) FROM userpoints_weekly LEFT JOIN points ON points.point_name = userpoints_weekly.point_name WHERE userpoints_weekly.telega_id = uw.telega_id GROUP BY telega_id) AS total_total_ratiosum,
    (SELECT SUM(point_score*ratio) FROM userpoints_weekly LEFT JOIN points ON points.point_name = userpoints_weekly.point_name WHERE userpoints_weekly.telega_id = uw.telega_id AND date BETWEEN ? AND ? GROUP BY telega_id) AS weekly_total_ratiosum,
    (SELECT SUM(userbonus_score) FROM user_bonus WHERE telega_id = uw.telega_id GROUP BY telega_id) AS last_bonuses_sum                                           
    FROM userpoints_weekly uw 
    LEFT JOIN usernames u ON uw.telega_id = u.telega_id 
    LEFT JOIN points p ON p.point_name = uw.point_name 
    WHERE date BETWEEN ? AND ?
    GROUP BY u.name, p.point_name
    ORDER BY weekly_total_ratiosum DESC
    """, (monday.date(), sunday.date(), monday.date(), sunday.date(), monday.date(), sunday.date(), monday.date(), sunday.date(), monday.date(), sunday.date(),))
    week_dates = [monday + timedelta(days=i) for i in range(7)]
    sqldata1 = sorted(sqldata1, key=lambda x: x['total_total_ratiosum'], reverse=True)
    sqldata2 = select_data_dict("SELECT usernames.name, userpoints_weekly.point_name, userpoints_weekly.date, userpoints_weekly.point_score, points.ratio FROM userpoints_weekly INNER JOIN usernames ON userpoints_weekly.telega_id = usernames.telega_id INNER JOIN points ON points.point_name = userpoints_weekly.point_name WHERE date BETWEEN ? AND ?", (monday.date(), sunday.date(), ))

    week_dates_str = [date.strftime('%Y-%m-%d') for date in week_dates]
    pred_result = []
    result = []

    for username in usernames:
        my_dict = {}
        my_dict['username'] = username
        my_dict['telega_id'] = 0 
        my_dict['Общая сумма'] = 0
        my_dict['Общая сумма с коэффициентом'] = 0
        my_dict['Бонус столбца'] = 0
        my_dict['Бонус строки'] = 0
        my_dict['Бонус минимумов'] = 0
        my_dict['Итог'] = 0
        for point in points:
            my_dict[point] = {}
            for week_day in week_dates_str:
                my_dict[point][week_day] = 0
            my_dict[point]["Недельная сумма поинта"] = 0
            my_dict[point]["Недельная сумма поинта с коэффициентом"] = 0
        pred_result.append(my_dict)

    for row in sqldata2:
        for result_row in pred_result:
            if result_row['username'] == row['name']:
                result_row[row['point_name']][row['date']] += row['point_score']
    
    for summs in sqldata1:
        for result_row in pred_result:
            if result_row['username'] == summs['name']:
                result_row[summs['point_name']]["Недельная сумма поинта"] = summs['point_week_sum']
                result_row[summs['point_name']]["Недельная сумма поинта с коэффициентом"] = summs['point_week_sum']*summs['ratio'] 
                result_row['Общая сумма'] += summs['point_week_sum']
                result_row['Общая сумма с коэффициентом'] = summs['weekly_total_ratiosum']
                result_row['Бонус столбца'] = bonus_ratio_editor(bonus=summs['bonus_of_cols'], my_num=summs['weekly_total_ratiosum'])*summs['count_column_bonus']
                result_row['Бонус строки'] = bonus_ratio_editor(bonus=summs['bonus_of_rows'], my_num=summs['weekly_total_ratiosum'])*summs['count_row_bonus']
                result_row['Бонус минимумов'] = bonus_ratio_editor(bonus=summs['bonus_of_mins'], my_num=summs['weekly_total_ratiosum'])*summs['bonus_of_mins_ratio']
                result_row['Итог'] = summs['weekly_total_ratiosum'] + bonus_ratio_editor(bonus=summs['bonus_of_cols'], my_num=summs['weekly_total_ratiosum'])*summs['count_column_bonus'] + bonus_ratio_editor(bonus=summs['bonus_of_rows'], my_num=summs['weekly_total_ratiosum'])*summs['count_row_bonus'] + bonus_ratio_editor(bonus=summs['bonus_of_mins'], my_num=summs['weekly_total_ratiosum'])*summs['bonus_of_mins_ratio']
                result_row['telega_id'] = summs['telega_id']                
    #sorted_data = sorted(data, key=lambda x: x['total_total_ratiosum'], reverse=True)
    pred_result =  sorted(pred_result, key=lambda x: x['Итог'], reverse=True) 
    for res in pred_result:
        heading = [res['username'], *week_dates_str, "Cумма поинта", "Cумма поинта с коэффициентом", "Бонус строки", "Бонус столбца", "Бонус минимумов", "Итог"]
        empty_list = ['']*7
        footer = ['Общее', *empty_list, res['Общая сумма'], res['Общая сумма с коэффициентом'], res['Бонус строки'], res['Бонус столбца'], res['Бонус минимумов'], res['Итог']]
        result.append(heading)
        for point in points:
            row = []
            row = [point, *res[point].values()]
            result.append(row)
        result.append([])
        result.append(footer)
        result.append([])
        result.append([])    
    return result

#print(womans_report(points=just_points, usernames=just_usenames, sqldata1=sqldata1))



def exsel_creator():
    just_points = select_data_single("SELECT point_name FROM points")
    just_usenames = select_data_single("SELECT name FROM usernames ORDER BY name")
    # Создаем новую рабочую книгу
    workbook = Workbook()
    #points_release = ["", *just_points, "Сумма", "Сумма c коэффициентом", "Сумма с бонусом"]
    # Добавляем новый лист
    sheet1 = workbook.active  # по умолчанию уже есть активный лист
    sheet1.title = "Подробный недельный отчет"  # переименовываем лист
    # Вставляем данные в Лист 1
    rprt = womans_report(points=just_points, usernames=just_usenames)
    for row in rprt:
        sheet1.append(row)
    for column in range(2,9):
        letter = utils.get_column_letter(column)
        sheet1.column_dimensions[letter].width = 100/9
    sheet1.column_dimensions['I'].width = 150/9
    sheet1.column_dimensions['J'].width = 210/9
    sheet1.column_dimensions['K'].width = 110/9
    sheet1.column_dimensions['L'].width = 120/9
    sheet1.column_dimensions['M'].width = 150/9
    # Создаем еще один лист
    sheet2 = workbook.create_sheet(title="Общий список")
    for x in ttl_ratiosum_process():
        sheet2.append([x['name'], x['total_ratiosun_with_bonuses']])        
    workbook.save("excels/Report.xlsx")

#exsel_creator(just_points, just_usenames)

